from __future__ import annotations

import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import argparse
import re
from dataclasses import dataclass
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from backend.settings.connection_points import DB_URL

@dataclass(frozen=True)
class ViewColumnRow:
    base_view_id: int
    column_id: int
    visible: bool
    editable: bool
    position: float

VIEW_RE = re.compile(r"View\s+(\d+)\s*-", re.IGNORECASE)

def parse_base_view_id(view_title: object) -> Optional[int]:
    if view_title is None:
        return None
    s = str(view_title).strip()
    if not s:
        return None
    m = VIEW_RE.search(s)
    return int(m.group(1)) if m else None

def read_headers_preview(
    xlsm_path: str,
    sheet_name: str = "Headers Preview",
    start_row: int = 2,
    start_col: int = 1,
    block_height: int = 5,
    block_spacing: int = 3,
    max_empty_blocks: int = 10,
    default_visible: bool = True,
    default_editable: bool = True,
) -> List[ViewColumnRow]:
    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found. Available: {wb.sheetnames}')
    ws = wb[sheet_name]
    step = block_height + block_spacing
    r = start_row
    empty_blocks = 0
    out: List[ViewColumnRow] = []
    while empty_blocks < max_empty_blocks:
        base_view_id_cell = ws.cell(row=r, column=start_col).value
        print(f"Row {r}, Col {start_col}, base_view_id: {base_view_id_cell}")  # Debug print
        if base_view_id_cell is None or str(base_view_id_cell).strip() == "":
            empty_blocks += 1
            r += step
            continue
        empty_blocks = 0
        try:
            base_view_id = int(str(base_view_id_cell).strip())
        except Exception as e:
            print(f"Skipping block at row {r}: cannot parse base_view_id from '{base_view_id_cell}'")  # Debug print
            r += step
            continue
        pos_row = r + 1
        id_row = r + 4
        c = start_col
        fallback_pos = 1
        while True:
            col_id_val = ws.cell(row=id_row, column=c).value
            print(f"  Col {c}: column_id={col_id_val}")  # Debug print
            if col_id_val is None or str(col_id_val).strip() == "":
                break
            try:
                column_id = int(str(col_id_val))
            except Exception as e:
                print(f"    Skipping col {c}: cannot parse column_id from '{col_id_val}'")  # Debug print
                break
            pos_val = ws.cell(row=pos_row, column=c).value
            print(f"    position={pos_val}")  # Debug print
            if pos_val is None or str(pos_val).strip() == "":
                position = float(fallback_pos)
            else:
                try:
                    position = float(str(pos_val))
                except Exception as e:
                    print(f"    Skipping col {c}: cannot parse position from '{pos_val}'")  # Debug print
                    break
            out.append(
                ViewColumnRow(
                    base_view_id=base_view_id,
                    column_id=column_id,
                    visible=default_visible,
                    editable=default_editable,
                    position=position,
                )
            )
            fallback_pos += 1
            c += 1
        r += step
    return out

UPSERT_SQL_TEMPLATE = """
INSERT INTO {table} (base_view_id, column_id, visible, editable, position)
VALUES %s
ON CONFLICT (base_view_id, column_id)
DO UPDATE SET
    visible = EXCLUDED.visible,
    editable = EXCLUDED.editable,
    position = EXCLUDED.position;
"""

def write_to_postgres(
    records: Sequence[ViewColumnRow],
    table: str,
    truncate_first: bool = False,
) -> None:
    try:
        import psycopg2
        from psycopg2.extras import execute_values
    except ImportError as e:
        raise SystemExit(
            "Missing dependency psycopg2.\n"
            "Install with:\n"
            "  pip install psycopg2-binary\n"
        ) from e
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    sql = UPSERT_SQL_TEMPLATE.format(table=table)
    values: List[Tuple[int, int, bool, bool, float]] = [
        (r.base_view_id, r.column_id, r.visible, r.editable, r.position) for r in records
    ]
    with conn:
        with conn.cursor() as cur:
            if truncate_first:
                cur.execute(f"TRUNCATE TABLE {table};")
            execute_values(cur, sql, values, page_size=1000)
    conn.close()

def add_missing_columns_from_excel(
    xlsm_path: str,
    sheet_name: str = "Headers Preview",
    start_row: int = 2,
    start_col: int = 1,
    block_height: int = 5,
    block_spacing: int = 3,
    max_empty_blocks: int = 10,
):
    import psycopg2
    wb = load_workbook(xlsm_path, data_only=True, keep_vba=True)
    ws = wb[sheet_name]
    step = block_height + block_spacing
    r = start_row
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = True
    with conn:
        with conn.cursor() as cur:
            while True:
                base_view_id_cell = ws.cell(row=r, column=start_col).value
                if base_view_id_cell is None or str(base_view_id_cell).strip() == "":
                    break
                pos_row = r + 1
                type_row = r + 3  # 4th row in block
                id_row = r + 4
                c = start_col
                while True:
                    col_id_val = ws.cell(row=id_row, column=c).value
                    if col_id_val is None or str(col_id_val).strip() == "":
                        break
                    try:
                        column_id = int(str(col_id_val))
                    except Exception:
                        c += 1
                        continue
                    type_val = ws.cell(row=type_row, column=c).value
                    type_digit = None
                    if type_val is not None and str(type_val).strip():
                        type_digit = str(type_val).strip()[0]
                    if type_digit and type_digit in "12345678":
                        # Get column name from columns table
                        cur.execute("SELECT name FROM columns WHERE id = %s", (column_id,))
                        row = cur.fetchone()
                        if not row:
                            c += 1
                            continue
                        col_name = row[0]
                        # Determine target tables
                        targets = []
                        if type_digit in "128":
                            targets = ["articles", "article_drafts"]
                        elif type_digit in "34567":
                            targets = ["project_articles"]
                        for table in targets:
                            # Check if column exists
                            cur.execute("""
                                SELECT 1 FROM information_schema.columns
                                WHERE table_name = %s AND column_name = %s
                            """, (table, col_name))
                            if not cur.fetchone():
                                # Add column as varchar (default type)
                                print(f"Adding column '{col_name}' to {table}")
                                cur.execute(f'ALTER TABLE {table} ADD COLUMN "{col_name}" VARCHAR')
                    c += 1
                r += step
    conn.close()

def generate_column_style_map_json(output_path: str = r"c:\Users\ferreres\my-ergonomic-table\src\frontend\visualization\Formating\ColumnStyleMap.json"):
    """
    Generate a JSON file mapping column group (color class) to color and headers, based on DB data.
    Structure matches ColumnStyleMap.json for frontend use.
    If the file exists, it will be overwritten to ensure a clean update.
    """
    import psycopg2
    import json
    import os
    # Remove the file if it exists to avoid merge/update issues
    if os.path.exists(output_path):
        os.remove(output_path)
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    try:
        cur.execute('''
            SELECT cg.color_name, cg.color, c.name_external_german
            FROM columns c
            LEFT JOIN column_groups cg ON c.column_group_id = cg.id
            WHERE c.name_external_german IS NOT NULL AND cg.color_name IS NOT NULL
        ''')
        style_map = {}
        for color_name, color, header in cur.fetchall():
            if color_name not in style_map:
                style_map[color_name] = {"color": color or "#f0f0f0", "headers": []}
            style_map[color_name]["headers"].append(header)
        # Add static entries for grid-header-rows and grid-inbetween-red (as in legacy)
        style_map["grid-header-rows"] = {
            "color": "#999999",
            "headers": [],
            "used_for": ["HEADER rows (Einbauort-Gruppen) Hintergrund"]
        }
        style_map["grid-inbetween-red"] = {
            "color": "#FF0000",
            "headers": [],
            "used_for": ["In-between '**' Markierung / Trennhinweis"]
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(style_map, f, ensure_ascii=False, indent=2)
        print(f"[INFO] Wrote column style map to {output_path}")
    finally:
        cur.close(); conn.close()

def generate_views_columns_auto_json(output_path: str = r"c:\Users\ferreres\my-ergonomic-table\src\frontend\visualization\views_columns_auto.json"):
    """
    Generate a JSON file with the ordered, visible columns for each base_view_id from views_columns_auto.
    Structure: { base_view_id: [ { column_id, position, visible, editable }, ... ] }
    """
    import psycopg2
    import json
    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    try:
        cur.execute('''
            SELECT base_view_id, column_id, visible, editable, position
            FROM views_columns_auto
            ORDER BY base_view_id, position
        ''')
        result = {}
        for base_view_id, column_id, visible, editable, position in cur.fetchall():
            if base_view_id not in result:
                result[base_view_id] = []
            result[base_view_id].append({
                "column_id": column_id,
                "visible": visible,
                "editable": editable,
                "position": position
            })
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"[INFO] Wrote views_columns_auto to {output_path}")
    finally:
        cur.close(); conn.close()

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Import Headers Preview -> views_columns_auto")
    p.add_argument(
        "--xlsm",
        default=r"C:\Users\ferreres\Documents\BachelorWSP\GeneralLayoutTemplate.xlsm",
        help="Path to GeneralLayout.xlsm",
    )
    p.add_argument("--sheet", default="Headers Preview", help="Sheet name to read")
    p.add_argument("--table", default="views_columns_auto", help="Target table")
    p.add_argument("--truncate", action="store_true", help="TRUNCATE table first")
    return p

def main() -> None:
    args = build_arg_parser().parse_args()
    if not os.path.exists(args.xlsm):
        raise SystemExit(f"Excel file not found: {args.xlsm}")
    records = read_headers_preview(args.xlsm, sheet_name=args.sheet)
    print(f"Parsed {len(records)} rows from Excel sheet '{args.sheet}'.")
    if not records:
        print("No records found. Nothing to write.")
        return
    write_to_postgres(
        records=records,
        table=args.table,
        truncate_first=args.truncate,
    )
    print(f"Upserted {len(records)} rows into table '{args.table}' (id is SERIAL4 in DB).")
    # Add missing columns to target tables
    add_missing_columns_from_excel(args.xlsm, sheet_name=args.sheet)
    # Generate the up-to-date ColumnStyleMap JSON
    generate_column_style_map_json()
    # No longer generate views_columns_auto JSON for frontend

if __name__ == "__main__":
    main()
